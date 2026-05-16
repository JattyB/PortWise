from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any


def write_excel_report(data: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return _write_minimal_xlsx(data, output_path)

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Executive Summary": _summary_rows(data),
        "Asset Inventory": _asset_rows(data),
        "Open Ports Summary": _port_rows(data),
        "Port-wise Service View": _service_rows(data),
        "Findings": _finding_rows(data.get("findings", [])),
        "Confirmed Findings": _finding_rows([f for f in data.get("findings", []) if f.get("confidence") == "Confirmed"]),
        "Needs Manual Validation": _finding_rows([f for f in data.get("findings", []) if f.get("manual_validation")]),
        "False Positive Candidates": _finding_rows([f for f in data.get("findings", []) if f.get("confidence") == "False Positive Candidate"]),
        "TLS Findings": _finding_rows([f for f in data.get("findings", []) if f.get("module") == "tls"]),
        "HTTP Findings": _finding_rows([f for f in data.get("findings", []) if f.get("module") == "http"]),
        "Exposure Findings": _finding_rows([f for f in data.get("findings", []) if f.get("module") == "exposure"]),
        "CVE Findings": _finding_rows([f for f in data.get("findings", []) if f.get("type") == "CVE"]),
        "UDP Findings": _udp_rows(data),
        "Module Targets": _module_target_rows(data),
        "Commands Executed": _command_rows(data),
        "Skipped and Failed Checks": _skipped_rows(data),
        "Retest Baseline": _finding_rows(data.get("findings", [])),
    }
    fills = {
        "critical": PatternFill("solid", fgColor="7F1D1D"),
        "high": PatternFill("solid", fgColor="FCA5A5"),
        "medium": PatternFill("solid", fgColor="FCD34D"),
        "low": PatternFill("solid", fgColor="BBF7D0"),
        "informational": PatternFill("solid", fgColor="E5E7EB"),
    }
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        for row in rows:
            ws.append([_cell_value(value) for value in row])
        if rows:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in ws.iter_rows(min_row=2):
                for item in row:
                    if str(item.value).lower() in fills:
                        item.fill = fills[str(item.value).lower()]
                    item.alignment = Alignment(wrap_text=True, vertical="top")
            for column_cells in ws.columns:
                width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 60)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    wb.save(output_path)
    return output_path


def _summary_rows(data: dict[str, Any]) -> list[list[Any]]:
    state = data.get("metadata", {}).get("state", {})
    findings = data.get("findings", [])
    rows = [["Metric", "Value"]]
    rows.extend([
        ["Total targets", len(state.get("targets_loaded", []))],
        ["Live hosts", len(state.get("live_hosts", []))],
        ["Services discovered", sum(len(v) for v in state.get("services_by_host", {}).values())],
        ["TCP ports found", sum(len(v) for v in state.get("tcp_open_ports_by_host", {}).values())],
        ["UDP ports found", sum(len(v) for v in state.get("udp_open_ports_by_host", {}).values())],
        ["Confirmed findings", sum(1 for f in findings if f.get("confidence") == "Confirmed")],
        ["Likely findings", sum(1 for f in findings if f.get("confidence") == "Likely")],
        ["Possible findings", sum(1 for f in findings if f.get("confidence") == "Possible")],
        ["Needs validation", sum(1 for f in findings if f.get("manual_validation"))],
    ])
    for priority in ("P1", "P2", "P3", "P4", "P5"):
        rows.append([priority, sum(1 for f in findings if f.get("priority") == priority)])
    return rows


def _finding_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    headers = ["Priority", "Status", "Severity", "Confidence", "Evidence Strength", "False Positive Risk", "Manual Validation", "Title", "Asset", "Port", "Protocol", "Module", "Type", "CVE", "Description", "Evidence", "Recommendation"]
    rows = [headers]
    for f in findings:
        rows.append([f.get("priority"), f.get("status"), f.get("severity"), f.get("confidence"), f.get("evidence_strength"), f.get("false_positive_risk"), f.get("manual_validation"), f.get("title"), f.get("asset"), f.get("port"), f.get("protocol"), f.get("module"), f.get("type"), f.get("cve_id"), f.get("description"), _short_evidence(f), f.get("recommendation")])
    return rows


def _asset_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows = [["Host", "Status", "Hostnames"]]
    for asset in data.get("assets", []):
        rows.append([asset.get("ip"), asset.get("status"), ", ".join(asset.get("hostnames", []))])
    return rows


def _service_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows = [["Host", "Port", "Protocol", "State", "Service", "Product", "Version", "CPE"]]
    for asset in data.get("assets", []):
        for svc in asset.get("services", []):
            rows.append([svc.get("host"), svc.get("port"), svc.get("protocol"), svc.get("state"), svc.get("service_name"), svc.get("product"), svc.get("version"), ", ".join(svc.get("cpes", []))])
    return rows


def _port_rows(data: dict[str, Any]) -> list[list[Any]]:
    state = data.get("metadata", {}).get("state", {})
    rows = [["Host", "TCP Open", "UDP Open", "UDP Open|Filtered"]]
    hosts = set(state.get("tcp_open_ports_by_host", {})) | set(state.get("udp_open_ports_by_host", {})) | set(state.get("udp_open_filtered_ports_by_host", {}))
    for host in sorted(hosts):
        rows.append([host, state.get("tcp_open_ports_by_host", {}).get(host, []), state.get("udp_open_ports_by_host", {}).get(host, []), state.get("udp_open_filtered_ports_by_host", {}).get(host, [])])
    return rows


def _udp_rows(data: dict[str, Any]) -> list[list[Any]]:
    state = data.get("metadata", {}).get("state", {})
    rows = [["Host", "UDP Open", "UDP Open|Filtered"]]
    for host in sorted(set(state.get("udp_open_ports_by_host", {})) | set(state.get("udp_open_filtered_ports_by_host", {}))):
        rows.append([host, state.get("udp_open_ports_by_host", {}).get(host, []), state.get("udp_open_filtered_ports_by_host", {}).get(host, [])])
    return rows


def _module_target_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows = [["Module Target List", "Host", "Port", "Protocol", "Service", "Reason"]]
    for key, targets in data.get("metadata", {}).get("state", {}).get("module_targets", {}).items():
        for t in targets:
            rows.append([key, t.get("host"), t.get("port"), t.get("protocol"), t.get("service"), t.get("routing_reason")])
    return rows


def _command_rows(data: dict[str, Any]) -> list[list[Any]]:
    rows = [["Name", "Command", "Skipped", "Return Code", "Error"]]
    for c in data.get("commands", []):
        rows.append([c.get("name"), " ".join(c.get("command", [])), c.get("skipped"), c.get("return_code"), c.get("error")])
    return rows


def _skipped_rows(data: dict[str, Any]) -> list[list[Any]]:
    state = data.get("metadata", {}).get("state", {})
    rows = [["Type", "Value"]]
    for item in state.get("skipped_phases", []) + data.get("skipped_checks", []):
        rows.append(["Skipped", item])
    for item in state.get("failed_phases", []) + data.get("failed_checks", []):
        rows.append(["Failed", item])
    return rows


def _short_evidence(finding: dict[str, Any]) -> str:
    evidence = finding.get("evidence", [])
    if not evidence:
        return ""
    return "; ".join(str(e.get("description", ""))[:180] for e in evidence[:3])


def _write_minimal_xlsx(data: dict[str, Any], output_path: Path) -> Path:
    # Valid single-sheet XLSX fallback for environments where openpyxl is not installed.
    rows = _summary_rows(data)
    sheet_rows = "".join("<row>" + "".join(f"<c t='inlineStr'><is><t>{_xml(str(cell))}</t></is></c>" for cell in row) + "</row>" for row in rows)
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "<?xml version='1.0'?><Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'><Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/><Default Extension='xml' ContentType='application/xml'/><Override PartName='/xl/workbook.xml' ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml'/><Override PartName='/xl/worksheets/sheet1.xml' ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'/></Types>")
        zf.writestr("_rels/.rels", "<?xml version='1.0'?><Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'><Relationship Id='rId1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument' Target='xl/workbook.xml'/></Relationships>")
        zf.writestr("xl/workbook.xml", "<?xml version='1.0'?><workbook xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main' xmlns:r='http://schemas.openxmlformats.org/officeDocument/2006/relationships'><sheets><sheet name='Executive Summary' sheetId='1' r:id='rId1'/></sheets></workbook>")
        zf.writestr("xl/_rels/workbook.xml.rels", "<?xml version='1.0'?><Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'><Relationship Id='rId1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet' Target='worksheets/sheet1.xml'/></Relationships>")
        zf.writestr("xl/worksheets/sheet1.xml", f"<?xml version='1.0'?><worksheet xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'><sheetData>{sheet_rows}</sheetData></worksheet>")
    return output_path


def _xml(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _cell_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)[:500]
    return value
